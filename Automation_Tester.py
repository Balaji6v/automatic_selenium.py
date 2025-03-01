import pytest
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

EXCEL_FILE = "test_data.xlsx"


# Function to create an Excel file if it doesn't exist
def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "LoginData"

        # Headers
        headers = ["Test ID", "Username", "Password", "Date Time", "Name of Tester", "Test Result"]
        sheet.append(headers)

        # Sample test data
        test_data = [
            [1, "Admin", "admin123", "2025-03-01 10:00", "Tester1", "Pending"],
            [2, "InvalidUser", "wrongpass", "2025-03-01 10:05", "Tester2", "Pending"],
            [3, "Admin", "admin123", "2025-03-01 10:10", "Tester3", "Pending"],
            [4, "TestUser", "test1234", "2025-03-01 10:15", "Tester4", "Pending"],
            [5, "Admin", "admin123", "2025-03-01 10:20", "Tester5", "Pending"]
        ]

        for row in test_data:
            sheet.append(row)

        workbook.save(EXCEL_FILE)
        print("✅ Excel file 'test_data.xlsx' created with sample data!")


# Load test data from Excel
def load_test_data():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return workbook, sheet, data


# Page Object Model (POM) for Login Page
class LoginPage:
    def __init__(self, driver):
        self.driver = driver
        self.username_field = (By.NAME, "username")
        self.password_field = (By.NAME, "password")
        self.login_button = (By.XPATH, "//button[@type='submit']")
        self.dashboard_text = (By.XPATH, "//span[text()='Dashboard']")

    def login(self, username, password):
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.username_field)).send_keys(username)
        WebDriverWait(self.driver, 10).until(
            EC.visibility_of_element_located(self.password_field)).send_keys(password)
        WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(self.login_button)).click()

    def is_login_successful(self):
        try:
            WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located(self.dashboard_text)
            )
            return True
        except:
            return False


# Ensure the Excel file is created before running tests
create_excel_file()

# Load test data from Excel
workbook, sheet, test_data = load_test_data()


@pytest.mark.parametrize("test_id, username, password, date_time, tester, result", test_data)
def test_login(test_id, username, password, date_time, tester, result):
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")

    login_page = LoginPage(driver)
    login_page.login(username, password)

    is_success = login_page.is_login_successful()

    # Update Excel with test result
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == test_id:
            row[5].value = "Pass" if is_success else "Fail"
            break

    # Save the updated Excel file
    workbook.save(EXCEL_FILE)

    # Close browser
    driver.quit()

